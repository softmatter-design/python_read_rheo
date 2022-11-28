#!/usr/bin/env python
# -*- coding: utf-8 -*-
#######################################################
import pandas as pd
import openpyxl
#######################################################

def read_book():
    data_file = 'datafile.xlsx'
    # input_book = openpyxl.load_workbook(data_file)
    # #sheet_namesメソッドでExcelブック内の各シートの名前をリストで取得できる
    # input_sheet_name = input_book.sheetnames
    # #lenでシートの総数を確認
    # num_sheet = len(input_sheet_name)
    # #シートの数とシートの名前のリストの表示
    # print ("Sheet の数:", num_sheet)
    # print (input_sheet_name)
    # target_sheet = input_book['data']
    # cell = target_sheet['A1']
    # print(cell.value)
    # g = target_sheet.iter_rows(min_row=1, max_row=30, min_col=1, max_col=10)
    # print(list)
    df = pd.read_excel(data_file, sheet_name=1, header=None, usecols=[0])
    print(df)
read_book()
